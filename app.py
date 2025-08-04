from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, make_response, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import os
from functools import wraps
import io
import base64
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///autoglass_mis.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Database Models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='staff')  # admin, staff, cashier
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)

class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    category = db.Column(db.String(50), nullable=False)  # glass, aluminum, accessories
    description = db.Column(db.Text)
    price = db.Column(db.Float, nullable=False)
    stock_quantity = db.Column(db.Integer, nullable=False, default=0)
    min_stock_level = db.Column(db.Integer, nullable=False, default=10)
    supplier_id = db.Column(db.Integer, db.ForeignKey('supplier.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Customer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    phone = db.Column(db.String(20))
    email = db.Column(db.String(120))
    address = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Supplier(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    contact_person = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    email = db.Column(db.String(120))
    address = db.Column(db.Text)
    products = db.relationship('Product', backref='supplier', lazy=True)
    purchase_orders = db.relationship('PurchaseOrder', backref='supplier', lazy=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Sale(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'))
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    total_amount = db.Column(db.Float, nullable=False)
    payment_method = db.Column(db.String(20), nullable=False)  # cash, credit
    status = db.Column(db.String(20), default='completed')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    items = db.relationship('SaleItem', backref='sale', lazy=True)

class SaleItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sale_id = db.Column(db.Integer, db.ForeignKey('sale.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    unit_price = db.Column(db.Float, nullable=False)
    total_price = db.Column(db.Float, nullable=False)
    product = db.relationship('Product', backref='sale_items')

class PurchaseOrder(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    po_number = db.Column(db.String(50), unique=True, nullable=False)
    supplier_id = db.Column(db.Integer, db.ForeignKey('supplier.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    total_amount = db.Column(db.Float, nullable=False)
    status = db.Column(db.String(20), default='pending')  # pending, approved, delivered, cancelled
    expected_delivery = db.Column(db.Date)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    items = db.relationship('PurchaseOrderItem', backref='purchase_order', lazy=True)

class PurchaseOrderItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    purchase_order_id = db.Column(db.Integer, db.ForeignKey('purchase_order.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    unit_price = db.Column(db.Float, nullable=False)
    total_price = db.Column(db.Float, nullable=False)
    product = db.relationship('Product', backref='purchase_order_items')

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Role-based access decorator
def role_required(role):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated or current_user.role != role:
                flash('Access denied. Insufficient permissions.', 'error')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Chart generation functions
def generate_sales_chart():
    """Generate sales chart for the last 30 days"""
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=30)
    
    # Get daily sales data
    sales_data = db.session.query(
        db.func.date(Sale.created_at).label('date'),
        db.func.sum(Sale.total_amount).label('total')
    ).filter(
        Sale.created_at >= start_date
    ).group_by(
        db.func.date(Sale.created_at)
    ).all()
    
    # Create chart
    plt.figure(figsize=(12, 6))
    dates = [item.date for item in sales_data]
    amounts = [float(item.total) for item in sales_data]
    
    plt.plot(dates, amounts, marker='o', linewidth=2, markersize=6)
    plt.title('Daily Sales - Last 30 Days', fontsize=16, fontweight='bold')
    plt.xlabel('Date', fontsize=12)
    plt.ylabel('Sales Amount ($)', fontsize=12)
    plt.xticks(rotation=45)
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    
    # Convert to base64 string
    img = io.BytesIO()
    plt.savefig(img, format='png', dpi=150, bbox_inches='tight')
    img.seek(0)
    chart_url = base64.b64encode(img.getvalue()).decode()
    plt.close()
    
    return chart_url

def generate_inventory_chart():
    """Generate inventory distribution chart by category"""
    # Get inventory data by category
    inventory_data = db.session.query(
        Product.category,
        db.func.count(Product.id).label('count'),
        db.func.sum(Product.stock_quantity * Product.price).label('value')
    ).group_by(Product.category).all()
    
    categories = [item.category.title() for item in inventory_data]
    values = [float(item.value) for item in inventory_data]
    
    # Create pie chart
    plt.figure(figsize=(10, 8))
    colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7']
    plt.pie(values, labels=categories, autopct='%1.1f%%', startangle=90, colors=colors)
    plt.title('Inventory Value Distribution by Category', fontsize=16, fontweight='bold')
    plt.axis('equal')
    
    # Convert to base64 string
    img = io.BytesIO()
    plt.savefig(img, format='png', dpi=150, bbox_inches='tight')
    img.seek(0)
    chart_url = base64.b64encode(img.getvalue()).decode()
    plt.close()
    
    return chart_url

# Routes
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    # Get dashboard statistics
    total_products = Product.query.count()
    low_stock_items = Product.query.filter(Product.stock_quantity <= Product.min_stock_level).count()
    today_sales = Sale.query.filter(Sale.created_at >= datetime.now().date()).count()
    total_customers = Customer.query.count()
    
    # Recent sales
    recent_sales = Sale.query.order_by(Sale.created_at.desc()).limit(5).all()
    
    return render_template('dashboard.html', 
                         total_products=total_products,
                         low_stock_items=low_stock_items,
                         today_sales=today_sales,
                         total_customers=total_customers,
                         recent_sales=recent_sales)

@app.route('/inventory')
@login_required
def inventory():
    search = request.args.get('search', '')
    category = request.args.get('category', '')
    
    query = Product.query
    if search:
        query = query.filter(Product.name.contains(search))
    if category:
        query = query.filter(Product.category == category)
    
    products = query.all()
    return render_template('inventory.html', products=products)

@app.route('/inventory/add', methods=['GET', 'POST'])
@login_required
def add_product():
    if request.method == 'POST':
        product = Product(
            name=request.form['name'],
            category=request.form['category'],
            description=request.form['description'],
            price=float(request.form['price']),
            stock_quantity=int(request.form['stock_quantity']),
            min_stock_level=int(request.form['min_stock_level']),
            supplier_id=request.form['supplier_id'] if request.form['supplier_id'] else None
        )
        db.session.add(product)
        db.session.commit()
        flash('Product added successfully!', 'success')
        return redirect(url_for('inventory'))
    
    suppliers = Supplier.query.all()
    return render_template('add_product.html', suppliers=suppliers)

@app.route('/sales')
@login_required
def sales():
    sales = Sale.query.order_by(Sale.created_at.desc()).all()
    return render_template('sales.html', sales=sales)

@app.route('/sales/new', methods=['GET', 'POST'])
@login_required
def new_sale():
    if request.method == 'POST':
        # Create new sale
        sale = Sale(
            customer_id=request.form['customer_id'] if request.form['customer_id'] else None,
            user_id=current_user.id,
            total_amount=0,
            payment_method=request.form['payment_method']
        )
        db.session.add(sale)
        db.session.flush()  # Get the sale ID
        
        total = 0
        # Add sale items
        for i in range(len(request.form.getlist('product_id'))):
            product_id = request.form.getlist('product_id')[i]
            quantity = int(request.form.getlist('quantity')[i])
            
            product = Product.query.get(product_id)
            unit_price = product.price
            total_price = unit_price * quantity
            total += total_price
            
            sale_item = SaleItem(
                sale_id=sale.id,
                product_id=product_id,
                quantity=quantity,
                unit_price=unit_price,
                total_price=total_price
            )
            db.session.add(sale_item)
            
            # Update stock
            product.stock_quantity -= quantity
        
        sale.total_amount = total
        db.session.commit()
        flash('Sale completed successfully!', 'success')
        return redirect(url_for('sales'))
    
    products = Product.query.filter(Product.stock_quantity > 0).all()
    customers = Customer.query.all()
    return render_template('new_sale.html', products=products, customers=customers)

@app.route('/customers')
@login_required
def customers():
    customers = Customer.query.all()
    return render_template('customers.html', customers=customers)

@app.route('/suppliers')
@login_required
def suppliers():
    suppliers = Supplier.query.all()
    return render_template('suppliers.html', suppliers=suppliers)

@app.route('/purchase-orders')
@login_required
def purchase_orders():
    orders = PurchaseOrder.query.order_by(PurchaseOrder.created_at.desc()).all()
    return render_template('purchase_orders.html', orders=orders)

@app.route('/purchase-orders/new', methods=['GET', 'POST'])
@login_required
def new_purchase_order():
    if request.method == 'POST':
        # Generate PO number
        po_count = PurchaseOrder.query.count() + 1
        po_number = f"PO-{po_count:04d}"
        
        # Create purchase order
        po = PurchaseOrder(
            po_number=po_number,
            supplier_id=request.form['supplier_id'],
            user_id=current_user.id,
            total_amount=0,
            expected_delivery=datetime.strptime(request.form['expected_delivery'], '%Y-%m-%d').date() if request.form['expected_delivery'] else None,
            notes=request.form['notes']
        )
        db.session.add(po)
        db.session.flush()
        
        total = 0
        # Add PO items
        for i in range(len(request.form.getlist('product_id'))):
            product_id = request.form.getlist('product_id')[i]
            quantity = int(request.form.getlist('quantity')[i])
            unit_price = float(request.form.getlist('unit_price')[i])
            total_price = unit_price * quantity
            total += total_price
            
            po_item = PurchaseOrderItem(
                purchase_order_id=po.id,
                product_id=product_id,
                quantity=quantity,
                unit_price=unit_price,
                total_price=total_price
            )
            db.session.add(po_item)
        
        po.total_amount = total
        db.session.commit()
        flash('Purchase order created successfully!', 'success')
        return redirect(url_for('purchase_orders'))
    
    suppliers = Supplier.query.all()
    products = Product.query.all()
    return render_template('new_purchase_order.html', suppliers=suppliers, products=products)

@app.route('/reports')
@login_required
@role_required('admin')
def reports():
    # Sales reports
    today = datetime.now().date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    
    daily_sales = db.session.query(db.func.sum(Sale.total_amount)).filter(
        Sale.created_at >= today
    ).scalar() or 0
    
    weekly_sales = db.session.query(db.func.sum(Sale.total_amount)).filter(
        Sale.created_at >= week_ago
    ).scalar() or 0
    
    monthly_sales = db.session.query(db.func.sum(Sale.total_amount)).filter(
        Sale.created_at >= month_ago
    ).scalar() or 0
    
    # Generate charts
    sales_chart = generate_sales_chart()
    inventory_chart = generate_inventory_chart()
    
    # Top selling products
    top_products = db.session.query(
        Product.name,
        db.func.sum(SaleItem.quantity).label('total_sold'),
        db.func.sum(SaleItem.total_price).label('total_revenue')
    ).join(SaleItem).group_by(Product.id).order_by(
        db.func.sum(SaleItem.quantity).desc()
    ).limit(5).all()
    
    return render_template('reports.html', 
                         daily_sales=daily_sales,
                         weekly_sales=weekly_sales,
                         monthly_sales=monthly_sales,
                         sales_chart=sales_chart,
                         inventory_chart=inventory_chart,
                         top_products=top_products)

@app.route('/export/sales-report')
@login_required
@role_required('admin')
def export_sales_report():
    format_type = request.args.get('format', 'pdf')
    
    # Get sales data
    sales = Sale.query.order_by(Sale.created_at.desc()).limit(100).all()
    
    if format_type == 'pdf':
        return generate_sales_pdf(sales)
    else:
        return generate_sales_excel(sales)

@app.route('/export/inventory-report')
@login_required
@role_required('admin')
def export_inventory_report():
    format_type = request.args.get('format', 'pdf')
    
    # Get inventory data
    products = Product.query.all()
    
    if format_type == 'pdf':
        return generate_inventory_pdf(products)
    else:
        return generate_inventory_excel(products)

@app.route('/export/purchase-order/<int:po_id>')
@login_required
def export_purchase_order(po_id):
    po = PurchaseOrder.query.get_or_404(po_id)
    return generate_purchase_order_pdf(po)

def generate_sales_pdf(sales):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title = Paragraph("Sales Report", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Date
    date_para = Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal'])
    story.append(date_para)
    story.append(Spacer(1, 12))
    
    # Table data
    data = [['Sale ID', 'Customer', 'Amount', 'Payment Method', 'Date']]
    for sale in sales:
        data.append([
            f"#{sale.id}",
            sale.customer.name if sale.customer else 'Walk-in',
            f"${sale.total_amount:.2f}",
            sale.payment_method.title(),
            sale.created_at.strftime('%Y-%m-%d')
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='sales_report.pdf', mimetype='application/pdf')

def generate_sales_excel(sales):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Headers
    headers = ['Sale ID', 'Customer', 'Amount', 'Payment Method', 'Date']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data
    for row, sale in enumerate(sales, 2):
        ws.cell(row=row, column=1, value=f"#{sale.id}")
        ws.cell(row=row, column=2, value=sale.customer.name if sale.customer else 'Walk-in')
        ws.cell(row=row, column=3, value=sale.total_amount)
        ws.cell(row=row, column=4, value=sale.payment_method.title())
        ws.cell(row=row, column=5, value=sale.created_at.strftime('%Y-%m-%d'))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, as_attachment=True, download_name='sales_report.xlsx', 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def generate_inventory_pdf(products):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title = Paragraph("Inventory Report", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Date
    date_para = Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal'])
    story.append(date_para)
    story.append(Spacer(1, 12))
    
    # Table data
    data = [['Product Name', 'Category', 'Stock', 'Price', 'Value']]
    total_value = 0
    for product in products:
        value = product.stock_quantity * product.price
        total_value += value
        data.append([
            product.name,
            product.category.title(),
            str(product.stock_quantity),
            f"${product.price:.2f}",
            f"${value:.2f}"
        ])
    
    # Add total row
    data.append(['', '', '', 'Total Value:', f"${total_value:.2f}"])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='inventory_report.pdf', mimetype='application/pdf')

def generate_inventory_excel(products):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    
    # Headers
    headers = ['Product Name', 'Category', 'Stock', 'Price', 'Total Value']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data
    total_value = 0
    for row, product in enumerate(products, 2):
        value = product.stock_quantity * product.price
        total_value += value
        ws.cell(row=row, column=1, value=product.name)
        ws.cell(row=row, column=2, value=product.category.title())
        ws.cell(row=row, column=3, value=product.stock_quantity)
        ws.cell(row=row, column=4, value=product.price)
        ws.cell(row=row, column=5, value=value)
    
    # Total row
    total_row = len(products) + 2
    ws.cell(row=total_row, column=4, value="Total Value:").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=total_value).font = Font(bold=True)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, as_attachment=True, download_name='inventory_report.xlsx', 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def generate_purchase_order_pdf(po):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Company header
    title = Paragraph("O. CASTRO AUTOGLASS & ALUMINUM", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 12))
    
    # PO Title
    po_title = Paragraph(f"PURCHASE ORDER - {po.po_number}", styles['Heading1'])
    story.append(po_title)
    story.append(Spacer(1, 12))
    
    # PO Details
    details = f"""
    <b>Supplier:</b> {po.supplier.name}<br/>
    <b>Date:</b> {po.created_at.strftime('%Y-%m-%d')}<br/>
    <b>Expected Delivery:</b> {po.expected_delivery.strftime('%Y-%m-%d') if po.expected_delivery else 'TBD'}<br/>
    <b>Status:</b> {po.status.title()}
    """
    details_para = Paragraph(details, styles['Normal'])
    story.append(details_para)
    story.append(Spacer(1, 20))
    
    # Items table
    data = [['Product', 'Quantity', 'Unit Price', 'Total']]
    for item in po.items:
        data.append([
            item.product.name,
            str(item.quantity),
            f"${item.unit_price:.2f}",
            f"${item.total_price:.2f}"
        ])
    
    # Total row
    data.append(['', '', 'TOTAL:', f"${po.total_amount:.2f}"])
    
    # Create table
    table = Table(data, colWidths=[3*inch, 1*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    
    if po.notes:
        story.append(Spacer(1, 20))
        notes_para = Paragraph(f"<b>Notes:</b> {po.notes}", styles['Normal'])
        story.append(notes_para)
    
    doc.build(story)
    
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f'PO_{po.po_number}.pdf', mimetype='application/pdf')

@app.route('/users')
@login_required
@role_required('admin')
def users():
    users = User.query.all()
    return render_template('users.html', users=users)

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        
        # Create default admin user if not exists
        if not User.query.filter_by(username='admin').first():
            admin = User(
                username='admin',
                email='admin@autoglass.com',
                password_hash=generate_password_hash('admin123'),
                role='admin'
            )
            db.session.add(admin)
            db.session.commit()
    
    app.run(debug=True)
